from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q
import json
from datetime import datetime
from .models import Car, Tracker, InstallationHistory, OrderDocument, Location
from .forms import CarForm, TrackerForm, InstallationForm

# Представление для карточек автомобилей
class CarListView(LoginRequiredMixin, ListView):
    model = Car
    template_name = 'tracker/car_list.html'
    context_object_name = 'cars'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Car.objects.all()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(board_number__icontains=search_query) |
                Q(model__icontains=search_query) |
                Q(location__icontains=search_query) |
                Q(comment__icontains=search_query)
            )
        # Sorting
        sort = self.request.GET.get('sort')
        allowed = {
            'state_number': 'state_number',
            'board_number': 'board_number',
            'model': 'model',
            'location': 'location__name',
            'is_active': 'is_active',
            'comment': 'comment',
            'created_at': 'created_at',
        }
        if sort:
            desc = sort.startswith('-')
            key = sort[1:] if desc else sort
            field = allowed.get(key)
            if field:
                queryset = queryset.order_by(f"{'-' if desc else ''}{field}")
        else:
            queryset = queryset.order_by('board_number')
        return queryset

class CarDetailView(LoginRequiredMixin, DetailView):
    model = Car
    template_name = 'tracker/car_detail.html'
    context_object_name = 'car'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['documents'] = self.object.order_documents.all()
        context['installations'] = self.object.installations.all().select_related('tracker')
        context['active_installation'] = self.object.installations.filter(is_active=True).first()
        return context

class CarCreateView(LoginRequiredMixin, CreateView):
    model = Car
    form_class = CarForm
    template_name = 'tracker/car_form.html'
    success_url = reverse_lazy('car_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Автомобиль успешно создан')
        return super().form_valid(form)

class CarUpdateView(LoginRequiredMixin, UpdateView):
    model = Car
    form_class = CarForm
    template_name = 'tracker/car_form.html'
    
    def get_success_url(self):
        return reverse_lazy('tracker:car_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['modal'] = (self.request.headers.get('x-requested-with') == 'XMLHttpRequest')
        return context
    
    def form_valid(self, form):
        messages.success(self.request, 'Изменения сохранены')
        response = super().form_valid(form)
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return HttpResponse('OK')
        return response

class CarDeleteView(LoginRequiredMixin, DeleteView):
    model = Car
    template_name = 'tracker/car_confirm_delete.html'
    success_url = reverse_lazy('car_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Автомобиль удален')
        return super().delete(request, *args, **kwargs)

# Представление для карточек трекеров (аналогично Car)
class TrackerListView(LoginRequiredMixin, ListView):
    model = Tracker
    template_name = 'tracker/tracker_list.html'
    context_object_name = 'trackers'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Tracker.objects.all()
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(serial_number__icontains=search_query) |
                Q(imei__icontains=search_query) |
                Q(model__icontains=search_query) |
                Q(comment__icontains=search_query)
            )
        sort = self.request.GET.get('sort')
        allowed = {
            'serial_number': 'serial_number',
            'imei': 'imei',
            'model': 'model',
            'protocol': 'protocol',
            'is_active': 'is_active',
            'created_at': 'created_at',
            'inventory_number_tracker': 'inventory_number_tracker',
            'inventory_number_antenna': 'inventory_number_antenna',
            'holder_number': 'holder_number',
            'sim_old': 'sim_old',
            'n_card': 'n_card',
            'sim_new': 'sim_new',
            'comment': 'comment',
            'location': 'installations__car__location__name',
        }
        if sort:
            desc = sort.startswith('-')
            key = sort[1:] if desc else sort
            field = allowed.get(key)
            if field:
                queryset = queryset.order_by(f"{'-' if desc else ''}{field}")
        else:
            queryset = queryset.order_by('serial_number')
        # prevent duplicate trackers due to joins on reverse FK when sorting by related fields
        return queryset.distinct()

# API для получения истории установок в JSON
def installation_history_api(request, car_id=None, tracker_id=None):
    """API для получения истории установок"""
    installations = InstallationHistory.objects.all()
    
    if car_id:
        installations = installations.filter(car_id=car_id)
    if tracker_id:
        installations = installations.filter(tracker_id=tracker_id)
    
    data = []
    for install in installations.select_related('car', 'tracker'):
        data.append({
            'id': install.id,
            'car': {
                'id': install.car.id,
                'board_number': install.car.board_number,
                'model': install.car.get_model_display(),
            },
            'tracker': {
                'id': install.tracker.id,
                'serial_number': install.tracker.serial_number,
                'imei': install.tracker.imei,
                'model': install.tracker.model,
            },
            'installation_date': install.installation_date.strftime('%Y-%m-%d'),
            'removal_date': install.removal_date.strftime('%Y-%m-%d') if install.removal_date else None,
            'is_active': install.is_active,
            'comment': install.comment,
        })
    
    return JsonResponse(data, safe=False)

# Представление для отчетов
class ReportView(LoginRequiredMixin, TemplateView):
    template_name = 'tracker/reports.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cars_count'] = Car.objects.count()
        context['trackers_count'] = Tracker.objects.count()
        context['active_installations'] = InstallationHistory.objects.filter(is_active=True).count()
        context['inactive_trackers'] = Tracker.objects.filter(
            is_active=True,
            installations__is_active=False
        ).distinct().count()

        # Filters from GET
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        location = self.request.GET.get('location')
        model_filter = self.request.GET.get('model')
        imei = self.request.GET.get('imei')
        serial = self.request.GET.get('serial')
        sim = self.request.GET.get('sim')

        qs = InstallationHistory.objects.select_related('car__location', 'tracker')
        # apply sorting from GET
        sort = self.request.GET.get('sort')
        allowed = {
            'car': 'car__board_number',
            'board_number': 'car__board_number',
            'state_number': 'car__state_number',
            'location': 'car__location__name',
            'tracker': 'tracker__serial_number',
            'installation_date': 'installation_date',
            'removal_date': 'removal_date',
            'is_active': 'is_active',
        }
        if sort:
            desc = sort.startswith('-')
            key = sort[1:] if desc else sort
            field = allowed.get(key)
            if field:
                qs = qs.order_by(f"{'-' if desc else ''}{field}")
        else:
            qs = qs.order_by('-installation_date')
        if date_from:
            qs = qs.filter(installation_date__gte=date_from)
        if date_to:
            qs = qs.filter(installation_date__lte=date_to)
        if location:
            try:
                loc_id = int(location)
                qs = qs.filter(car__location__id=loc_id)
            except ValueError:
                qs = qs.filter(car__location__name=location)
        if model_filter:
            qs = qs.filter(car__model=model_filter)
        if imei:
            qs = qs.filter(tracker__imei__icontains=imei)
        if serial:
            qs = qs.filter(tracker__serial_number__icontains=serial)
        if sim:
            qs = qs.filter(Q(tracker__sim_old__icontains=sim) | Q(tracker__sim_new__icontains=sim) | Q(tracker__n_card__icontains=sim))

        context['installations'] = qs[:100]
        from .models import Location
        context['location_choices'] = Location.objects.all()
        # provide models choices from Car.CAR_MODELS
        context['model_choices'] = Car.CAR_MODELS
        context['filters'] = {'date_from': date_from or '', 'date_to': date_to or '', 'location': location or '', 'model': model_filter or '', 'imei': imei or '', 'serial': serial or '', 'sim': sim or ''}
        return context


# Export installations to XLSX
from django.utils import timezone
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

from django.http import HttpResponse

@login_required
def export_installations_xlsx(request):
    qs = InstallationHistory.objects.select_related('car__location', 'tracker').order_by('-installation_date')

    # Support filters from GET (same as reports)
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    location = request.GET.get('location')
    model_filter = request.GET.get('model')
    imei = request.GET.get('imei')
    serial = request.GET.get('serial')
    sim = request.GET.get('sim')

    if date_from:
        qs = qs.filter(installation_date__gte=date_from)
    if date_to:
        qs = qs.filter(installation_date__lte=date_to)
    if location:
        try:
            loc_id = int(location)
            qs = qs.filter(car__location__id=loc_id)
        except ValueError:
            qs = qs.filter(car__location__name=location)
    if model_filter:
        qs = qs.filter(car__model=model_filter)
    if imei:
        qs = qs.filter(tracker__imei__icontains=imei)
    if serial:
        qs = qs.filter(tracker__serial_number__icontains=serial)
    if sim:
        qs = qs.filter(Q(tracker__sim_old__icontains=sim) | Q(tracker__sim_new__icontains=sim) | Q(tracker__n_card__icontains=sim))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Installations'

    from openpyxl.styles import PatternFill

    headers = ['Бортовой номер', 'Державний номер', 'Локация', 'Трекер (S/N)', 'Дата установки', 'Дата снятия', 'Активна', 'Комментарий']
    ws.append(headers)

    # header fill
    header_fill = PatternFill(fill_type='solid', fgColor='FFDDEAF6')
    for c in ws[1]:
        c.fill = header_fill

    # Write rows and apply zebra fill
    odd_fill = PatternFill(fill_type='solid', fgColor='FFFFFFFF')
    even_fill = PatternFill(fill_type='solid', fgColor='FFEDF7FF')
    row_idx = 2
    for inst in qs:
        loc_value = inst.car.location.name if inst.car.location else '-'
        row = [
            inst.car.board_number or '-',
            inst.car.state_number or '-',
            loc_value,
            inst.tracker.serial_number,
            inst.installation_date.strftime('%Y-%m-%d'),
            inst.removal_date.strftime('%Y-%m-%d') if inst.removal_date else '',
            'Да' if inst.is_active else 'Нет',
            inst.comment or '',
        ]
        ws.append(row)
        # apply fill to the newly appended row
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for cell in ws[row_idx]:
            cell.fill = fill
        row_idx += 1

    # Auto width
    for i, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = min(50, length + 2)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"installations_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class TrackerDetailView(LoginRequiredMixin, DetailView):
    model = Tracker
    template_name = 'tracker/tracker_detail.html'
    context_object_name = 'tracker'


class TrackerCreateView(LoginRequiredMixin, CreateView):
    model = Tracker
    form_class = TrackerForm
    template_name = 'tracker/tracker_form.html'
    success_url = reverse_lazy('tracker:tracker_list')

    def form_valid(self, form):
        messages.success(self.request, 'Трекер успешно создан')
        response = super().form_valid(form)
        # If a car was selected during creation, create an active installation
        current_car = form.cleaned_data.get('current_car', None)
        if current_car:
            today = timezone.now().date()
            InstallationHistory.objects.create(car=current_car, tracker=self.object, installation_date=today, is_active=True, comment='Назначено из формы трекера')
        return response


class TrackerUpdateView(LoginRequiredMixin, UpdateView):
    model = Tracker
    form_class = TrackerForm
    template_name = 'tracker/tracker_form.html'

    def get_success_url(self):
        return reverse_lazy('tracker:tracker_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['modal'] = (self.request.headers.get('x-requested-with') == 'XMLHttpRequest')
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Изменения сохранены')
        # capture selection before save
        selected_car = form.cleaned_data.get('current_car', None)
        response = super().form_valid(form)
        # Apply changes to installations if needed
        today = timezone.now().date()
        active = self.object.installations.filter(is_active=True).first()
        # assign to a car
        if selected_car:
            if not active or active.car_id != selected_car.id:
                if active:
                    active.is_active = False
                    active.removal_date = today
                    active.save()
                InstallationHistory.objects.create(car=selected_car, tracker=self.object, installation_date=today, is_active=True, comment='Назначено из формы трекера')
        else:
            # clearing assignment: deactivate any active installation
            if active:
                active.is_active = False
                active.removal_date = today
                active.save()
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return HttpResponse('OK')
        return response


class TrackerDeleteView(LoginRequiredMixin, DeleteView):
    model = Tracker
    template_name = 'tracker/tracker_confirm_delete.html'
    success_url = reverse_lazy('tracker:tracker_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Трекер удалён')
        return super().delete(request, *args, **kwargs)


class InstallationListView(LoginRequiredMixin, ListView):
    model = InstallationHistory
    template_name = 'tracker/installation_list.html'
    context_object_name = 'installations'
    paginate_by = 20

    def get_queryset(self):
        queryset = InstallationHistory.objects.select_related('car', 'tracker')
        sort = self.request.GET.get('sort')
        allowed = {
            'car': 'car__board_number',
            'tracker': 'tracker__serial_number',
            'tracker_location': 'tracker__installations__car__location__name',
            'installation_date': 'installation_date',
            'removal_date': 'removal_date',
            'is_active': 'is_active',
        }
        if sort:
            desc = sort.startswith('-')
            key = sort[1:] if desc else sort
            field = allowed.get(key)
            if field:
                queryset = queryset.order_by(f"{'-' if desc else ''}{field}")
        else:
            queryset = queryset.order_by('-installation_date')
        return queryset

    def get_context_data(self, **kwargs):
        """Expose whether any OrderDocument exists so the template can show a global prompt only when needed."""
        context = super().get_context_data(**kwargs)
        from .models import OrderDocument
        context['order_docs_exist'] = OrderDocument.objects.exists()
        return context


class InstallationCreateView(LoginRequiredMixin, CreateView):
    model = InstallationHistory
    form_class = InstallationForm
    template_name = 'tracker/installation_form.html'
    success_url = reverse_lazy('tracker:installation_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import OrderDocument
        context['order_docs_exist'] = OrderDocument.objects.exists()
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Установка успешно создана')
        return super().form_valid(form)


class InstallationUpdateView(LoginRequiredMixin, UpdateView):
    model = InstallationHistory
    form_class = InstallationForm
    template_name = 'tracker/installation_form.html'

    def get_success_url(self):
        return reverse_lazy('tracker:installation_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['modal'] = (self.request.headers.get('x-requested-with') == 'XMLHttpRequest')
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Изменения сохранены')
        response = super().form_valid(form)
        if self.request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return HttpResponse('OK')
        return response


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'tracker/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Support sorting on dashboard recent installations via ?sort=
        sort = self.request.GET.get('sort')
        allowed = {
            'imei': 'tracker__imei',
            'board_number': 'car__board_number',
            'model': 'car__model',
            'protocol': 'tracker__protocol',
            'state_number': 'car__state_number',
            'sim_old': 'tracker__sim_old',
            'n_card': 'tracker__n_card',
            'sim_new': 'tracker__sim_new',
            'serial_number': 'tracker__serial_number',
            'inventory_number_tracker': 'tracker__inventory_number_tracker',
            'inventory_number_antenna': 'tracker__inventory_number_antenna',
            'installation_date': 'installation_date',
            'comment': 'comment',
        }
        qs = InstallationHistory.objects.select_related('car', 'tracker')
        if sort:
            desc = sort.startswith('-')
            key = sort[1:] if desc else sort
            field = allowed.get(key)
            if field:
                qs = qs.order_by(f"{'-' if desc else ''}{field}")
        else:
            qs = qs.order_by('-installation_date')
        context['recent_installations'] = qs[:10]
        context['cars_count'] = Car.objects.count()
        context['trackers_count'] = Tracker.objects.count()
        return context