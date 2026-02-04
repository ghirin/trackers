from django.test import TestCase, Client
from django.contrib.auth import get_user_model
from django.urls import reverse
from tracker.models import Car, Tracker, InstallationHistory, Location
from datetime import date

class SortingTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.user = User.objects.create_user('u','u@example.com','pw')
        self.client = Client()
        self.client.login(username='u', password='pw')
        self.loc = Location.objects.create(name='L1')
        self.loc2 = Location.objects.create(name='L2')
        self.car1 = Car.objects.create(board_number='B1', state_number='S1', model='Other', location=self.loc)
        self.car2 = Car.objects.create(board_number='B2', state_number='S2', model='Other', location=self.loc)
        self.car3 = Car.objects.create(board_number='B3', state_number='S3', model='Other', location=self.loc2)
        self.tr1 = Tracker.objects.create(
            serial_number='SN2', imei='123456789012345', model='ModelX', inventory_number_tracker='1110000000',
            inventory_number_antenna='1110000000', protocol='wialon', holder_number='АМ2000АВ', sim_old='4596734596', n_card='N111', sim_new='4596730000', comment='Test tracker 1'
        )
        self.tr2 = Tracker.objects.create(
            serial_number='SN1', imei='223456789012345', model='ModelY', inventory_number_tracker='1110000001',
            inventory_number_antenna='1110000002', protocol='wialon', holder_number='АМ2000АВ', sim_old='4596734597', n_card='N112', sim_new='4596730001', comment='Test tracker 2'
        )
        # installations with dates
        self.inst1 = InstallationHistory.objects.create(car=self.car1, tracker=self.tr1, installation_date=date(2021,1,1), is_active=True)
        self.inst2 = InstallationHistory.objects.create(car=self.car2, tracker=self.tr2, installation_date=date(2022,1,1), is_active=False)
        # active installation for tr2 on car3 so tracker locations differ
        self.inst3 = InstallationHistory.objects.create(car=self.car3, tracker=self.tr2, installation_date=date(2023,1,1), is_active=True)

    def test_car_sort_by_state_number_asc_desc(self):
        url = reverse('tracker:car_list')
        r = self.client.get(url + '?sort=state_number')
        self.assertEqual(r.status_code, 200)
        cars = list(r.context['cars'])
        self.assertEqual(cars[0].state_number, 'S1')
        r2 = self.client.get(url + '?sort=-state_number')
        cars2 = list(r2.context['cars'])
        self.assertEqual(cars2[0].state_number, 'S3')

    def test_tracker_sort_by_serial(self):
        url = reverse('tracker:tracker_list')
        r = self.client.get(url + '?sort=serial_number')
        trackers = list(r.context['trackers'])
        self.assertEqual(trackers[0].serial_number, 'SN1')

    def test_installation_sort_by_date(self):
        url = reverse('tracker:installation_list')
        r = self.client.get(url + '?sort=installation_date')
        inst = list(r.context['installations'])
        self.assertEqual(inst[0].installation_date, self.inst1.installation_date)
        r2 = self.client.get(url + '?sort=-installation_date')
        inst2 = list(r2.context['installations'])
        # newest installation is inst3 (2023)
        self.assertEqual(inst2[0].installation_date, self.inst3.installation_date)

    def test_dashboard_sort_by_imei(self):
        url = reverse('tracker:dashboard')
        r = self.client.get(url + '?sort=imei')
        recent = list(r.context['recent_installations'])
        self.assertEqual(recent[0].tracker.imei, '123456789012345')
        r2 = self.client.get(url + '?sort=-imei')
        recent2 = list(r2.context['recent_installations'])
        self.assertEqual(recent2[0].tracker.imei, '223456789012345')

    def test_dashboard_sort_by_state_number(self):
        url = reverse('tracker:dashboard')
        r = self.client.get(url + '?sort=state_number')
        recent = list(r.context['recent_installations'])
        self.assertEqual(recent[0].car.state_number, 'S1')
        r2 = self.client.get(url + '?sort=-state_number')
        recent2 = list(r2.context['recent_installations'])
        self.assertEqual(recent2[0].car.state_number, 'S3')
    def test_reports_sort_by_car(self):
        url = reverse('tracker:reports')
        r = self.client.get(url + '?sort=car')
        inst = list(r.context['installations'])
        # car board numbers sorted B1, B2
        self.assertEqual(inst[0].car.board_number, 'B1')

    def test_reports_sort_by_board_number(self):
        url = reverse('tracker:reports')
        r = self.client.get(url + '?sort=board_number')
        inst = list(r.context['installations'])
        self.assertEqual(inst[0].car.board_number, 'B1')
        r2 = self.client.get(url + '?sort=-board_number')
        inst2 = list(r2.context['installations'])
        self.assertEqual(inst2[0].car.board_number, 'B3')

    def test_reports_sort_by_state_number(self):
        url = reverse('tracker:reports')
        r = self.client.get(url + '?sort=state_number')
        inst = list(r.context['installations'])
        self.assertEqual(inst[0].car.state_number, 'S1')
        r2 = self.client.get(url + '?sort=-state_number')
        inst2 = list(r2.context['installations'])
        self.assertEqual(inst2[0].car.state_number, 'S3')

    def test_export_xlsx_headers_and_striping(self):
        url = reverse('tracker:export_installations')
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        # load workbook from response content
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        # headers should contain separate columns
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn('Бортовой номер', headers)
        self.assertIn('Державний номер', headers)
        # check zebra striping: row 2 and row 3 should not have same fill color (if present)
        row2_fill = ws[2][0].fill.fgColor.value
        row3_fill = ws[3][0].fill.fgColor.value
        self.assertNotEqual(row2_fill, row3_fill)

    def test_tracker_sort_by_location(self):
        url = reverse('tracker:tracker_list')
        r = self.client.get(url + '?sort=location')
        trackers = list(r.context['trackers'])
        # trackers should be sorted by current location name (L1 then L2)
        self.assertEqual(trackers[0].current_location, 'L1')
        r2 = self.client.get(url + '?sort=-location')
        trackers2 = list(r2.context['trackers'])
        self.assertEqual(trackers2[0].current_location, 'L2')

    def test_tracker_headers_tooltips_and_active_position(self):
        url = reverse('tracker:tracker_list')
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200)
        content = r.content.decode('utf-8')
        # tooltip attributes present
        self.assertIn('data-bs-toggle="tooltip"', content)
        # titles for headers (check a couple)
        self.assertIn('title="Сортировать по IMEI"', content)
        self.assertIn('title="Сортировать по активности"', content)
        # ensure 'Комментарий' header comes before 'Активен'
        self.assertTrue(content.find('Комментарий') < content.find('Активен'))
        # active status is displayed as checkbox
        self.assertIn('type="checkbox"', content)

    def test_tracker_status_symbols_active_and_inactive(self):
        # make one tracker inactive and ensure we render cross for inactive and checkbox for active
        self.tr2.is_active = False
        self.tr2.save()
        url = reverse('tracker:tracker_list')
        r = self.client.get(url)
        content = r.content.decode('utf-8')
        self.assertIn('✖', content)
        # there's at least one checkbox for active trackers
        self.assertIn('type="checkbox"', content)

    def test_installation_comment_tooltip_when_long(self):
        long_comment = 'x' * 200
        from datetime import date
        InstallationHistory.objects.create(car=self.car1, tracker=self.tr1, installation_date=date(2024,1,1), comment=long_comment, is_active=False)
        url = reverse('tracker:installation_list')
        r = self.client.get(url)
        content = r.content.decode('utf-8')
        # tooltip attribute and full comment are rendered
        self.assertIn('data-bs-toggle="tooltip"', content)
        self.assertIn(long_comment, content)

    def test_installation_sort_by_tracker_location(self):
        url = reverse('tracker:installation_list')
        r = self.client.get(url + '?sort=tracker_location')
        inst = list(r.context['installations'])
        # installation list ordered by tracker's location
        self.assertEqual(inst[0].tracker.current_location, 'L1')
        r2 = self.client.get(url + '?sort=-tracker_location')
        inst2 = list(r2.context['installations'])
        self.assertEqual(inst2[0].tracker.current_location, 'L2')